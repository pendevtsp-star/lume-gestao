from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def as_text(value):
    if value is None:
        return ""
    return str(value)


def xlsx_response(filename, sheets):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="EDE6D8")
    header_font = Font(bold=True, color="202624")

    for title, headers, rows in sheets:
        worksheet = workbook.create_sheet(title[:31])
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            worksheet.append([as_text(value) for value in row])

        for column_cells in worksheet.columns:
            max_length = max(len(as_text(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def pdf_response(filename, title, sections=None, tables=None, landscape_page=False):
    buffer = BytesIO()
    page_size = landscape(A4) if landscape_page else A4
    margin = 1.25 * cm
    document = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=margin,
        bottomMargin=1.15 * cm,
    )
    styles = _pdf_styles()
    story = [_build_pdf_header(title, page_size, margin, styles), Spacer(1, 0.28 * cm)]

    for heading, lines in sections or []:
        story.append(_build_section_card(heading, lines, page_size, margin, styles))
        story.append(Spacer(1, 0.22 * cm))

    for heading, headers, rows in tables or []:
        story.append(Paragraph(heading, styles["SectionTitle"]))
        story.append(Spacer(1, 0.08 * cm))
        table_data = [[Paragraph(as_text(value), styles["TableHeader"]) for value in headers]]
        for row in rows:
            table_data.append([Paragraph(as_text(value), styles["TableCell"]) for value in row])

        col_widths = _table_column_widths(headers, rows, page_size[0] - (margin * 2))
        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEDC")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#263028")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#AEBB91")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2D8C6")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBF8F0")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.32 * cm))

    document.build(story, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _pdf_styles():
    base = getSampleStyleSheet()
    base.add(
        ParagraphStyle(
            name="ReportTitle",
            parent=base["Title"],
            alignment=0,
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#263028"),
            spaceAfter=4,
        )
    )
    base.add(
        ParagraphStyle(
            name="ReportMeta",
            parent=base["BodyText"],
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#6E7469"),
        )
    )
    base.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11.5,
            leading=14,
            textColor=colors.HexColor("#60724F"),
            spaceAfter=2,
        )
    )
    base.add(
        ParagraphStyle(
            name="SectionLine",
            parent=base["BodyText"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#263028"),
        )
    )
    base.add(
        ParagraphStyle(
            name="TableHeader",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.2,
            leading=10,
            textColor=colors.HexColor("#263028"),
        )
    )
    base.add(
        ParagraphStyle(
            name="TableCell",
            parent=base["BodyText"],
            fontSize=8.2,
            leading=10.5,
            textColor=colors.HexColor("#263028"),
            wordWrap="CJK",
        )
    )
    return base


def _build_pdf_header(title, page_size, margin, styles):
    available_width = page_size[0] - (margin * 2)
    logo = _logo_flowable()
    left_column = []
    if logo:
        left_column.append(logo)
    else:
        left_column.append(Paragraph("Lume", styles["ReportTitle"]))

    generated_at = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    right_column = [
        Paragraph(title, styles["ReportTitle"]),
        Paragraph(f"Lume Gestao - gerado em {generated_at}", styles["ReportMeta"]),
    ]
    table = Table([[left_column, right_column]], colWidths=[4.8 * cm, available_width - 4.8 * cm])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FBF8F0")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#DED4BD")),
                ("LINEBELOW", (0, 0), (-1, -1), 1.4, colors.HexColor("#AEBB91")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table


def _build_section_card(heading, lines, page_size, margin, styles):
    available_width = page_size[0] - (margin * 2)
    content = [Paragraph(heading, styles["SectionTitle"])]
    for line in lines:
        content.append(Paragraph(as_text(line), styles["SectionLine"]))
    table = Table([[content]], colWidths=[available_width])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFDF7")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2D8C6")),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _table_column_widths(headers, rows, available_width):
    if not headers:
        return None
    weights = []
    for index, header in enumerate(headers):
        values = [as_text(header)] + [as_text(row[index]) for row in rows[:30] if index < len(row)]
        longest = max((len(value) for value in values), default=8)
        weights.append(min(max(longest, 8), 28))
    total = sum(weights) or len(headers)
    min_width = 1.75 * cm
    widths = [max(min_width, available_width * (weight / total)) for weight in weights]
    width_total = sum(widths)
    if width_total > available_width:
        scale = available_width / width_total
        widths = [width * scale for width in widths]
    return widths


def _draw_pdf_footer(canvas, document):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#DED4BD"))
    canvas.setLineWidth(0.4)
    canvas.line(document.leftMargin, 0.78 * cm, document.pagesize[0] - document.rightMargin, 0.78 * cm)
    canvas.setFillColor(colors.HexColor("#6E7469"))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(document.leftMargin, 0.45 * cm, "Lume Gestao")
    canvas.drawRightString(document.pagesize[0] - document.rightMargin, 0.45 * cm, f"Pagina {document.page}")
    canvas.restoreState()


def _logo_path():
    candidates = [
        Path(settings.MEDIA_ROOT) / "clinic" / "IMG_3887.PNG",
        Path(settings.BASE_DIR) / "desktop" / "build" / "icon.png",
        Path(settings.BASE_DIR) / "desktop" / "build" / "icon-512.png",
    ]
    for path in candidates:
        if path.exists():
            return path
    return None


def _logo_flowable():
    logo_path = _logo_path()
    if not logo_path:
        return None
    try:
        from PIL import Image as PILImage

        image = PILImage.open(logo_path)
        image.thumbnail((520, 292))
        buffer = BytesIO()
        image.save(buffer, format="PNG", optimize=True)
        buffer.seek(0)
        return Image(buffer, width=4.4 * cm, height=2.45 * cm, kind="proportional")
    except Exception:
        return Image(str(logo_path), width=4.4 * cm, height=2.45 * cm, kind="proportional")
